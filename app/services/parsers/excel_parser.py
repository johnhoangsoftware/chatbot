"""
Excel file parser for .xlsx and .xls files.
"""

import os
from typing import List
from .base import BaseParser, ParsedDocument, TableData

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls)."""
    
    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xls']
    
    def parse(self, file_path: str) -> ParsedDocument:
        """Parse an Excel file."""
        if not self.validate(file_path):
            raise ValueError(f"Invalid or unsupported file: {file_path}")
        
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.xlsx':
            return self._parse_xlsx(file_path)
        elif ext == '.xls':
            return self._parse_xls(file_path)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")
    
    def _parse_xlsx(self, file_path: str) -> ParsedDocument:
        """Parse .xlsx file using openpyxl."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for .xlsx files. Install with: pip install openpyxl")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        full_content = []
        pages = []
        all_tables = []
        page_num = 1
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Extract sheet data
            sheet_content = []
            sheet_content.append(f"=== Sheet: {sheet_name} ===\n")
            
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                rows_data.append(row_text)
                sheet_content.append(row_text)
            
            # Create table data for this sheet
            if rows_data:
                table_content = "\n".join(rows_data)
                all_tables.append(TableData(
                    page_number=page_num,
                    content=table_content,
                    rows=len(rows_data),
                    columns=sheet.max_column or 0
                ))
            
            sheet_text = "\n".join(sheet_content)
            full_content.append(sheet_text)
            
            # Treat each sheet as a "page"
            pages.append({
                "page_number": page_num,
                "content": sheet_text,
                "word_count": len(sheet_text.split()),
                "char_count": len(sheet_text),
                "has_tables": True,
                "has_figures": False,
                "sections": [sheet_name]
            })
            page_num += 1
        
        # Build metadata
        metadata = {
            "filename": os.path.basename(file_path),
            "file_path": file_path,
            "page_count": len(pages),
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "total_words": sum(p["word_count"] for p in pages),
            "total_chars": sum(p["char_count"] for p in pages),
            "table_count": len(all_tables),
            "section_count": 0,
            "has_toc": False,
            "document_type": "Excel Spreadsheet"
        }
        
        workbook.close()
        
        return ParsedDocument(
            filename=os.path.basename(file_path),
            content="\n\n".join(full_content),
            pages=pages,
            metadata=metadata,
            sections=[],
            tables=all_tables,
            toc=[]
        )
    
    def _parse_xls(self, file_path: str) -> ParsedDocument:
        """Parse .xls file using xlrd."""
        if not XLRD_AVAILABLE:
            raise ImportError("xlrd is required for .xls files. Install with: pip install xlrd")
        
        workbook = xlrd.open_workbook(file_path)
        
        full_content = []
        pages = []
        all_tables = []
        page_num = 1
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            
            # Extract sheet data
            sheet_content = []
            sheet_content.append(f"=== Sheet: {sheet_name} ===\n")
            
            rows_data = []
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                
                # Skip empty rows
                if all(str(cell).strip() == '' for cell in row):
                    continue
                
                row_text = " | ".join(str(cell) for cell in row)
                rows_data.append(row_text)
                sheet_content.append(row_text)
            
            # Create table data for this sheet
            if rows_data:
                table_content = "\n".join(rows_data)
                all_tables.append(TableData(
                    page_number=page_num,
                    content=table_content,
                    rows=len(rows_data),
                    columns=sheet.ncols
                ))
            
            sheet_text = "\n".join(sheet_content)
            full_content.append(sheet_text)
            
            # Treat each sheet as a "page"
            pages.append({
                "page_number": page_num,
                "content": sheet_text,
                "word_count": len(sheet_text.split()),
                "char_count": len(sheet_text),
                "has_tables": True,
                "has_figures": False,
                "sections": [sheet_name]
            })
            page_num += 1
        
        # Build metadata
        sheet_names = [workbook.sheet_by_index(i).name for i in range(workbook.nsheets)]
        metadata = {
            "filename": os.path.basename(file_path),
            "file_path": file_path,
            "page_count": len(pages),
            "sheet_count": workbook.nsheets,
            "sheet_names": sheet_names,
            "total_words": sum(p["word_count"] for p in pages),
            "total_chars": sum(p["char_count"] for p in pages),
            "table_count": len(all_tables),
            "section_count": 0,
            "has_toc": False,
            "document_type": "Excel Spreadsheet (Legacy)"
        }
        
        return ParsedDocument(
            filename=os.path.basename(file_path),
            content="\n\n".join(full_content),
            pages=pages,
            metadata=metadata,
            sections=[],
            tables=all_tables,
            toc=[]
        )
